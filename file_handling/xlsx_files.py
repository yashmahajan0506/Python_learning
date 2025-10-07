from openpyxl import Workbook

wb=Workbook()

sheet=wb.active

sheet.title="yashh_file"


data=[
    ['id','name','age','designation'],
    [1,'yash',20,'CEO'],
    [2,'sujal',30,'CFO']
]

for i in data:
    sheet.append(i)
    
wb.save("ytt.xlsx")
# print("The data is saved")
# from openpyxl import load_workbook

# wb=load_workbook('yash.xlsx')

# sheet=wb.active

# # sheet.append([3,'kush',22,'Accounted'])
# # wb.save("yash.xlsx")
# for i in sheet.iter_rows(values_only=True):
#     print(i)
