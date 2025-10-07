from openpyxl import Workbook

wb=Workbook()

sheet=wb.active

sheet.title='employee'

data = [
    ["ID", "Name", "Department", "Salary"],
    [1, "Yash", "HR", 50000],
    [2, "Sujal", "Sales", 64000],
    [3, "Piyush", "IT", 70340],
    [4, "Dev", "Finance", 11000],
 ]

for i in data:
    sheet.append(i)
wb.save("file_excel.xlsx")
print("Data saved successfuly:")


#Fro printing Data:-  

# from openpyxl import load_workbook  

# workbook=load_workbook('excel_file.xlsx')

# sheet=workbook.active

# for i in sheet.iter_rows(values_only=True):
#     print(i)

